"""
The first test case.

Check creation of a proper xlsx file
"""

import datetime
import pathlib
import unittest

import openpyxl

# TODO: move it into __init__.py
CWD = pathlib.Path(__file__).resolve().parent


# TODO: rename according to its behavior
class TestCreateBasicWorkbook(unittest.TestCase):
    """A test case that creates and fills up a spreadsheet."""

    @classmethod
    def setUpClass(cls):
        """FIXME: refactor cls.cwd."""
        cls.cwd = pathlib.Path(__file__).resolve().parent

    def setUp(self):
        """Find out absolute path to the working directory.
        hint: this function is called before each test method.
        setUpClass is called once before all test methods"""
        pass

    def test_creation(self):
        """Make an object of class Workbook."""
        wb = openpyxl.Workbook()  # FIXME: self.wb?
        ws1 = wb.active  # Get active worksheet

        # Fill up some cells
        ws1["A1"], ws1["B1"], ws1["C1"] = 1, 2, 3
        ws1.append([4, 5, 6])  # Append cells starting from next line?
        ws1.append([7, 8, 9])  # Append cells starting from next line?

        ws1.title = "Matrix"

        # All other workbook / worksheet attributes are NOT copied - e.g.
        # Images, Charts.
        if "Matrix" in wb.sheetnames:
            _ws1 = wb.copy_worksheet(wb["Matrix"])
            _ws1.sheet_properties.tabColor = "1072BA"
            _ws1.title = "Matrix hidden copy"

            cells = _ws1["A1":"C3"]
            self.assertEqual(
                sum(cell.value for column in cells for cell in column),
                sum(range(1, 9 + 1)),  # Sum from 1 to 9
            )

        # Create and fill up second sheet
        ws2 = wb.create_sheet("Main Report", 0)
        ws2["A2"] = datetime.datetime.now()
        ws2.column_dimensions["A"].width = 20.0

        # Save on a disk
        # FIXME: use module defined variable
        dest_path = self.cwd / "output" / "1st_test_create_workbook.xlsx"
        try:
            wb.save(dest_path)
        except PermissionError as err:
            # https://stackoverflow.com/questions/41910583/errno-13-permission-denied-python
            if err.args[0] == 13:
                self.fail("File is already opened in other application.\
                           Close it!")
                # No need to raise Exception cause `fail` method
                # raises failureException which is AssertionError

    def test_loading(self):
        """Load xlsx created with Excel to make sure everything is parsed
        correctly."""
        pass

    def tearDown(self):
        """"""
        pass
